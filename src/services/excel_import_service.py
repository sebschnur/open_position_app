"""Service: Excel-Mockdatenimport.

Vorgabe: docs/specifications/06_excel_mockdaten_import.md. Liest die dort
definierten Zellbereiche aus der Mockdatei und uebersetzt die alten
Excel-Richtungen (K/V, Kauf/Verkauf) in die neue Vorzeichenlogik der App.
Die Excel-Datei dient nur der initialen Mockbefuellung; danach arbeitet die
App ausschliesslich gegen SQLite.

Bewusste Entscheidungen (siehe Projekt-Memory / bestaetigte Klaerpunkte):
- Position: Netto je Jahr = Summe(V-Mengen) - Summe(K-Mengen), keine
  Default-Fallback-Werte bei fehlenden Jahren (Position ist fachlich zu
  wichtig fuer stillschweigend erfundene Werte - stattdessen nur Warnung).
- Limitorder: nur "Kauf-Limit" wird unterstuetzt (-> partner_buys_price_lt_limit,
  siehe bestaetigter Klaerpunkt zu `<=`). "Verkauf-Limit" wird ignoriert und
  gemeldet. Verantwortlicher Handel/Vertrieb fehlen in der Excel-Datei und
  werden als klar gekennzeichneter Default-Platzhalter gesetzt.
- Settlementpreise: nur Base kommt aus Excel, Peak ist immer ein Default.
- Kein automatisches Setzen von Status "abgelaufen" bei bereits verstrichenem
  `valid_until` - dieser Status ist laut Spezifikation noch nicht geklaert
  (04_workflows_automatisierungen.md, Abschnitt 16).
"""

import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy.orm import Session

from src.db import default_mock_values as _defaults
from src.db.models import (
    IntradayTrade,
    LimitOrder,
    MarketPrice,
    OtcSurcharge,
    PfcCheck,
    PortfolioPosition,
    SettlementPrice,
    TradingCalendarEntry,
)
from src.domain.validation import (
    DIRECTION_PARTNER_BUYS,
    DIRECTION_PARTNER_SELLS,
    at_least_one_nonzero,
    expected_sign_for_direction,
    expected_sign_for_trigger_condition,
    validate_trade_quantities,
)
from src.repositories.metadata_repository import upsert_metadata
from src.services.initialization_service import is_database_seeded
from src.services.pfc_import_service import read_pfc_checks_from_files

SOURCE_LABEL = "excel_mock_file"

_POSITION_SHEET = "Position Frontjahre"
_PRICES_SHEET = "Vertriebsinfos"
_CALENDAR_SHEET = "Handelskalender"
_LIMIT_ORDER_SHEET = "Limitorder"

_POSITION_YEAR_HEADER_ROW = 6
_POSITION_YEAR_COLS = range(4, 9)  # D..H
_POSITION_DATA_ROWS = (7, 8)  # Zeile 7 = K, Zeile 8 = V
_INTRADAY_DATA_START_ROW = 9

_PRICE_YEAR_COL = 1  # Spalte A
_PRICE_DATA_ROWS = range(14, 18)
_SETTLEMENT_COL = 12  # Spalte L

_CALENDAR_HEADER_ROW = 2
_CALENDAR_YEAR_COLS = range(5, 8)  # E, F, G
_CALENDAR_DATA_START_ROW = 3

_LIMIT_ORDER_DATA_START_ROW = 5

_YEAR_OFFSETS_ALL = range(5)  # Y0..Y4
_YEAR_OFFSETS_PRICES = (1, 2, 3)  # Y+1..Y+3


@dataclass
class PortfolioPositionSeed:
    year: int
    position_mwh: float


@dataclass
class IntradayTradeSeed:
    trade_date: dt.date
    partner_alias: str
    quantity_y0_mwh: float
    quantity_y1_mwh: float
    quantity_y2_mwh: float
    quantity_y3_mwh: float
    quantity_y4_mwh: float


@dataclass
class MarketPriceSeed:
    product_type: str
    delivery_year: int
    price_eur_mwh: float


@dataclass
class SurchargeSeed:
    product_type: str
    delivery_year: int
    surcharge_eur_mwh: float


@dataclass
class SettlementPriceSeed:
    product_type: str
    delivery_year: int
    settlement_price_eur_mwh: float
    is_default: bool


@dataclass
class TradingCalendarSeed:
    due_date: dt.date
    partner_alias: str
    direction: str
    quantity_y0_mwh: float
    quantity_y1_mwh: float
    quantity_y2_mwh: float
    quantity_y3_mwh: float
    quantity_y4_mwh: float


@dataclass
class LimitOrderSeed:
    partner_alias: str
    quantity_y0_mwh: float
    quantity_y1_mwh: float
    quantity_y2_mwh: float
    quantity_y3_mwh: float
    quantity_y4_mwh: float
    trigger_product_type: str
    trigger_delivery_year: int
    trigger_condition: str
    limit_price_eur_mwh: float
    responsible_trading: str
    responsible_sales: str
    valid_until: Optional[dt.date]


@dataclass
class ImportReport:
    warnings: List[str]
    already_seeded: bool


def _year_offsets_from_header(ws, header_row: int, cols, current_year: int) -> Dict[int, int]:
    """Liefert {Spalte: Offset} fuer Header-Spalten, deren Jahr zu Y0..Y4 passt."""
    mapping: Dict[int, int] = {}
    for col in cols:
        year = ws.cell(row=header_row, column=col).value
        if isinstance(year, float) and year.is_integer():
            year = int(year)
        if not isinstance(year, int):
            continue
        offset = year - current_year
        if 0 <= offset <= 4:
            mapping[col] = offset
    return mapping


def _sign_from_kv(value) -> Optional[float]:
    if not isinstance(value, str):
        return None
    normalized = value.strip().upper()
    if normalized == "V":
        return 1.0
    if normalized == "K":
        return -1.0
    return None


def _direction_from_excel(value) -> Optional[str]:
    if not isinstance(value, str):
        return None
    normalized = value.strip().lower()
    if normalized == "kauf":
        return DIRECTION_PARTNER_BUYS
    if normalized == "verkauf":
        return DIRECTION_PARTNER_SELLS
    return None


def _cell_date(value, fallback: dt.date) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return fallback


def read_portfolio_positions_from_excel(
    workbook, current_year: int
) -> Tuple[List[PortfolioPositionSeed], List[str]]:
    """Liest die Netto-PMS-Position je Jahr (Summe(V) - Summe(K))."""
    warnings: List[str] = []
    ws = workbook[_POSITION_SHEET]
    col_offsets = _year_offsets_from_header(
        ws, _POSITION_YEAR_HEADER_ROW, _POSITION_YEAR_COLS, current_year
    )
    if not col_offsets:
        warnings.append(
            f"'{_POSITION_SHEET}': Keine Jahresspalten passen zu Y0..Y4 "
            f"({current_year}..{current_year + 4}) - Position kann nicht importiert werden."
        )

    net_by_offset: Dict[int, float] = {}
    for row in _POSITION_DATA_ROWS:
        sign = _sign_from_kv(ws.cell(row=row, column=3).value)
        if sign is None:
            warnings.append(
                f"'{_POSITION_SHEET}' Zeile {row}: unbekannte Richtung, Zeile wird ignoriert."
            )
            continue
        for col, offset in col_offsets.items():
            value = ws.cell(row=row, column=col).value
            if isinstance(value, (int, float)):
                net_by_offset[offset] = net_by_offset.get(offset, 0.0) + sign * float(value)

    seeds: List[PortfolioPositionSeed] = []
    for offset in _YEAR_OFFSETS_ALL:
        year = current_year + offset
        if offset in net_by_offset:
            seeds.append(PortfolioPositionSeed(year=year, position_mwh=net_by_offset[offset]))
        else:
            # Bewusst KEIN Default-Fallback: die PMS-Position ist fachlich zu
            # wichtig (Limitpruefung), um sie stillschweigend zu erfinden.
            warnings.append(
                f"Keine PMS-Position fuer Jahr {year} (Y+{offset}) in Excel gefunden - "
                "wird nicht importiert (kein Default fuer Positionen)."
            )
    return seeds, warnings


def read_intraday_trades_from_excel(
    workbook, current_year: int, today: dt.date
) -> Tuple[List[IntradayTradeSeed], List[str]]:
    """Liest untertaegige Geschaefte (je Excel-Zeile ein Geschaeft, K/V bestimmt Vorzeichen)."""
    warnings: List[str] = []
    ws = workbook[_POSITION_SHEET]
    col_offsets = _year_offsets_from_header(
        ws, _POSITION_YEAR_HEADER_ROW, _POSITION_YEAR_COLS, current_year
    )

    seeds: List[IntradayTradeSeed] = []
    for row in range(_INTRADAY_DATA_START_ROW, ws.max_row + 1):
        partner_alias = ws.cell(row=row, column=2).value
        if not partner_alias:
            continue

        sign = _sign_from_kv(ws.cell(row=row, column=3).value)
        if sign is None:
            warnings.append(
                f"'{_POSITION_SHEET}' Zeile {row} ('{partner_alias}'): unbekannte Richtung, "
                "Zeile wird ignoriert."
            )
            continue

        quantities = {offset: 0.0 for offset in _YEAR_OFFSETS_ALL}
        for col, offset in col_offsets.items():
            value = ws.cell(row=row, column=col).value
            if isinstance(value, (int, float)):
                quantities[offset] = sign * float(value)

        if not at_least_one_nonzero(quantities.values()):
            warnings.append(
                f"'{_POSITION_SHEET}' Zeile {row} ('{partner_alias}'): alle Mengen sind 0, "
                "Zeile wird ignoriert."
            )
            continue

        trade_date = _cell_date(ws.cell(row=row, column=1).value, fallback=today)
        seeds.append(
            IntradayTradeSeed(
                trade_date=trade_date,
                partner_alias=str(partner_alias).strip(),
                quantity_y0_mwh=quantities[0],
                quantity_y1_mwh=quantities[1],
                quantity_y2_mwh=quantities[2],
                quantity_y3_mwh=quantities[3],
                quantity_y4_mwh=quantities[4],
            )
        )
    return seeds, warnings


def _read_price_and_surcharge_block(
    ws, product_type: str, price_col: int, surcharge_col: int, current_year: int
) -> Tuple[List[MarketPriceSeed], List[SurchargeSeed], List[str]]:
    warnings: List[str] = []
    market_prices: List[MarketPriceSeed] = []
    surcharges: List[SurchargeSeed] = []
    found_offsets = set()

    for row in _PRICE_DATA_ROWS:
        year = ws.cell(row=row, column=_PRICE_YEAR_COL).value
        if not isinstance(year, int):
            continue
        offset = year - current_year
        if offset not in _YEAR_OFFSETS_PRICES:
            continue
        found_offsets.add(offset)

        price = ws.cell(row=row, column=price_col).value
        if isinstance(price, (int, float)):
            market_prices.append(
                MarketPriceSeed(product_type=product_type, delivery_year=year, price_eur_mwh=float(price))
            )
        else:
            warnings.append(
                f"'{_PRICES_SHEET}' Zeile {row}: Marktpreis {product_type} {year} fehlt/ist nicht "
                "numerisch - wird nicht importiert."
            )

        surcharge = ws.cell(row=row, column=surcharge_col).value
        if isinstance(surcharge, (int, float)):
            surcharges.append(
                SurchargeSeed(product_type=product_type, delivery_year=year, surcharge_eur_mwh=float(surcharge))
            )
        else:
            warnings.append(
                f"'{_PRICES_SHEET}' Zeile {row}: OTC-Aufschlag {product_type} {year} fehlt/ist nicht "
                "numerisch - wird nicht importiert."
            )

    for offset in _YEAR_OFFSETS_PRICES:
        if offset not in found_offsets:
            warnings.append(
                f"'{_PRICES_SHEET}': Keine Zeile fuer {product_type} Y+{offset} "
                f"({current_year + offset}) gefunden."
            )

    return market_prices, surcharges, warnings


def read_market_prices_and_surcharges_from_excel(
    workbook, current_year: int
) -> Tuple[List[MarketPriceSeed], List[SurchargeSeed], List[str]]:
    """Liest Marktpreise/OTC-Aufschlaege fuer Base und Peak, Y+1 bis Y+3."""
    ws = workbook[_PRICES_SHEET]
    base_prices, base_surcharges, base_warnings = _read_price_and_surcharge_block(
        ws, "Base", price_col=2, surcharge_col=3, current_year=current_year
    )
    peak_prices, peak_surcharges, peak_warnings = _read_price_and_surcharge_block(
        ws, "Peak", price_col=7, surcharge_col=8, current_year=current_year
    )
    return (
        base_prices + peak_prices,
        base_surcharges + peak_surcharges,
        base_warnings + peak_warnings,
    )


def read_settlement_prices_from_excel(
    workbook, current_year: int
) -> Tuple[List[SettlementPriceSeed], List[str]]:
    """Liest Base-Settlementpreise aus Excel; Peak ist immer ein Default (siehe Spec 8.1)."""
    warnings: List[str] = []
    ws = workbook[_PRICES_SHEET]
    seeds: List[SettlementPriceSeed] = []
    found_offsets = set()

    for row in _PRICE_DATA_ROWS:
        year = ws.cell(row=row, column=_PRICE_YEAR_COL).value
        if not isinstance(year, int):
            continue
        offset = year - current_year
        if offset not in _YEAR_OFFSETS_PRICES:
            continue

        value = ws.cell(row=row, column=_SETTLEMENT_COL).value
        if isinstance(value, (int, float)):
            seeds.append(
                SettlementPriceSeed(
                    product_type="Base", delivery_year=year,
                    settlement_price_eur_mwh=float(value), is_default=False,
                )
            )
            found_offsets.add(offset)
        else:
            warnings.append(
                f"'{_PRICES_SHEET}' Zeile {row}: Settlementpreis Base {year} fehlt/ist nicht "
                "numerisch - Default-Mockwert wird verwendet."
            )

    for offset in _YEAR_OFFSETS_PRICES:
        if offset not in found_offsets:
            year = current_year + offset
            seeds.append(
                SettlementPriceSeed(
                    product_type="Base", delivery_year=year,
                    settlement_price_eur_mwh=_defaults.BASE_SETTLEMENT_PRICE[offset], is_default=True,
                )
            )

    warnings.append(
        "Peak-Settlementpreise sind nicht in der Excel-Mockdatei enthalten - "
        "Default-Mockwerte werden verwendet."
    )
    for offset in _YEAR_OFFSETS_PRICES:
        year = current_year + offset
        seeds.append(
            SettlementPriceSeed(
                product_type="Peak", delivery_year=year,
                settlement_price_eur_mwh=_defaults.PEAK_SETTLEMENT_PRICE[offset], is_default=True,
            )
        )

    return seeds, warnings


def read_trading_calendar_from_excel(
    workbook, current_year: int, today: dt.date
) -> Tuple[List[TradingCalendarSeed], List[str]]:
    """Liest Handelskalendereintraege (Spalten A:G), Produktspalte D wird nicht uebernommen."""
    warnings: List[str] = []
    ws = workbook[_CALENDAR_SHEET]
    col_offsets = _year_offsets_from_header(
        ws, _CALENDAR_HEADER_ROW, _CALENDAR_YEAR_COLS, current_year
    )
    if not col_offsets:
        warnings.append(
            f"'{_CALENDAR_SHEET}': Keine Jahresspalten passen zu Y+1..Y+3 "
            f"({current_year + 1}..{current_year + 3})."
        )

    seeds: List[TradingCalendarSeed] = []
    for row in range(_CALENDAR_DATA_START_ROW, ws.max_row + 1):
        partner_alias = ws.cell(row=row, column=2).value
        if not partner_alias:
            continue

        direction = _direction_from_excel(ws.cell(row=row, column=3).value)
        if direction is None:
            warnings.append(
                f"'{_CALENDAR_SHEET}' Zeile {row} ('{partner_alias}'): unbekannte Richtung, "
                "Eintrag wird ignoriert."
            )
            continue

        sign = 1.0 if direction == DIRECTION_PARTNER_BUYS else -1.0
        quantities = {offset: 0.0 for offset in _YEAR_OFFSETS_ALL}
        for col, offset in col_offsets.items():
            value = ws.cell(row=row, column=col).value
            if isinstance(value, (int, float)):
                quantities[offset] = sign * abs(float(value))

        quantities_list = [quantities[offset] for offset in _YEAR_OFFSETS_ALL]
        errors = validate_trade_quantities(quantities_list, expected_sign_for_direction(direction))
        if errors:
            warnings.append(
                f"'{_CALENDAR_SHEET}' Zeile {row} ('{partner_alias}'): {' '.join(errors)} "
                "Eintrag wird ignoriert."
            )
            continue

        due_date = _cell_date(ws.cell(row=row, column=1).value, fallback=today)
        seeds.append(
            TradingCalendarSeed(
                due_date=due_date,
                partner_alias=str(partner_alias).strip(),
                direction=direction,
                quantity_y0_mwh=quantities[0],
                quantity_y1_mwh=quantities[1],
                quantity_y2_mwh=quantities[2],
                quantity_y3_mwh=quantities[3],
                quantity_y4_mwh=quantities[4],
            )
        )
    return seeds, warnings


def read_limit_orders_from_excel(
    workbook, current_year: int
) -> Tuple[List[LimitOrderSeed], List[str]]:
    """Liest Limitorders; nur 'Kauf-Limit' wird unterstuetzt (siehe Modulkommentar)."""
    warnings: List[str] = []
    ws = workbook[_LIMIT_ORDER_SHEET]
    seeds: List[LimitOrderSeed] = []

    for row in range(_LIMIT_ORDER_DATA_START_ROW, ws.max_row + 1):
        delivery_year = ws.cell(row=row, column=1).value
        partner_alias = ws.cell(row=row, column=2).value
        if not isinstance(delivery_year, int) or not partner_alias:
            continue

        quantity = ws.cell(row=row, column=3).value
        product_type = ws.cell(row=row, column=4).value
        buy_limit = ws.cell(row=row, column=5).value
        sell_limit = ws.cell(row=row, column=6).value
        valid_until_value = ws.cell(row=row, column=7).value

        if product_type not in ("Base", "Peak"):
            warnings.append(
                f"'{_LIMIT_ORDER_SHEET}' Zeile {row} ('{partner_alias}'): unbekanntes Produkt "
                f"'{product_type}', Eintrag wird ignoriert."
            )
            continue

        if not isinstance(quantity, (int, float)) or quantity == 0:
            warnings.append(
                f"'{_LIMIT_ORDER_SHEET}' Zeile {row} ('{partner_alias}'): Menge fehlt/ist 0, "
                "Eintrag wird ignoriert."
            )
            continue

        if isinstance(buy_limit, (int, float)):
            trigger_condition = "partner_buys_price_lt_limit"
            limit_price = float(buy_limit)
        elif isinstance(sell_limit, (int, float)):
            warnings.append(
                f"'{_LIMIT_ORDER_SHEET}' Zeile {row} ('{partner_alias}'): Verkauf-Limit wird im "
                "Mock-Import nicht unterstuetzt, Eintrag wird ignoriert."
            )
            continue
        else:
            warnings.append(
                f"'{_LIMIT_ORDER_SHEET}' Zeile {row} ('{partner_alias}'): weder Kauf- noch "
                "Verkauf-Limit befuellt, Eintrag wird ignoriert."
            )
            continue

        offset = delivery_year - current_year
        if not (0 <= offset <= 4):
            warnings.append(
                f"'{_LIMIT_ORDER_SHEET}' Zeile {row} ('{partner_alias}'): Lieferjahr {delivery_year} "
                f"liegt ausserhalb Y0..Y4 ({current_year}..{current_year + 4}), Eintrag wird ignoriert."
            )
            continue

        quantities = {o: 0.0 for o in _YEAR_OFFSETS_ALL}
        quantities[offset] = abs(float(quantity))
        quantities_list = [quantities[o] for o in _YEAR_OFFSETS_ALL]
        errors = validate_trade_quantities(
            quantities_list, expected_sign_for_trigger_condition(trigger_condition)
        )
        if errors:
            warnings.append(
                f"'{_LIMIT_ORDER_SHEET}' Zeile {row} ('{partner_alias}'): {' '.join(errors)} "
                "Eintrag wird ignoriert."
            )
            continue

        warnings.append(
            f"'{_LIMIT_ORDER_SHEET}' Zeile {row} ('{partner_alias}'): Verantwortlicher Handel/Vertrieb "
            "nicht in Excel vorhanden, Default-Platzhalter wird verwendet."
        )

        valid_until = None
        if isinstance(valid_until_value, dt.datetime):
            valid_until = valid_until_value.date()
        elif isinstance(valid_until_value, dt.date):
            valid_until = valid_until_value

        seeds.append(
            LimitOrderSeed(
                partner_alias=str(partner_alias).strip(),
                quantity_y0_mwh=quantities[0],
                quantity_y1_mwh=quantities[1],
                quantity_y2_mwh=quantities[2],
                quantity_y3_mwh=quantities[3],
                quantity_y4_mwh=quantities[4],
                trigger_product_type=product_type,
                trigger_delivery_year=delivery_year,
                trigger_condition=trigger_condition,
                limit_price_eur_mwh=limit_price,
                responsible_trading="Handel (Default Import)",
                responsible_sales="Vertrieb (Default Import)",
                valid_until=valid_until,
            )
        )
    return seeds, warnings


def seed_database_from_excel(
    session: Session,
    excel_path: Path,
    pfc_dir: Optional[Path] = None,
    today: Optional[dt.date] = None,
) -> ImportReport:
    """Befuellt eine leere Datenbank aus der Excel-Mockdatei (und optional PFC-Dateien).

    Ueberschreibt eine bereits initialisierte Datenbank NICHT (siehe
    06_excel_mockdaten_import.md, Abschnitt 1.1). Committet nicht selbst -
    das entscheidet der Aufrufer (analog zu seed_default_mock_data).
    """
    if is_database_seeded(session):
        return ImportReport(warnings=[], already_seeded=True)

    today = today or dt.date.today()
    now = dt.datetime.utcnow()
    current_year = today.year
    all_warnings: List[str] = []

    workbook = openpyxl.load_workbook(excel_path, data_only=True)

    position_seeds, w = read_portfolio_positions_from_excel(workbook, current_year)
    all_warnings += w
    for seed in position_seeds:
        session.add(
            PortfolioPosition(
                as_of_date=today, year=seed.year, position_mwh=seed.position_mwh, source=SOURCE_LABEL
            )
        )

    trade_seeds, w = read_intraday_trades_from_excel(workbook, current_year, today)
    all_warnings += w
    for seed in trade_seeds:
        session.add(
            IntradayTrade(
                trade_date=seed.trade_date,
                partner_alias=seed.partner_alias,
                quantity_y0_mwh=seed.quantity_y0_mwh,
                quantity_y1_mwh=seed.quantity_y1_mwh,
                quantity_y2_mwh=seed.quantity_y2_mwh,
                quantity_y3_mwh=seed.quantity_y3_mwh,
                quantity_y4_mwh=seed.quantity_y4_mwh,
                source_type="manual",
            )
        )

    price_seeds, surcharge_seeds, w = read_market_prices_and_surcharges_from_excel(workbook, current_year)
    all_warnings += w
    for seed in price_seeds:
        session.add(
            MarketPrice(
                product_type=seed.product_type, delivery_year=seed.delivery_year,
                price_eur_mwh=seed.price_eur_mwh, price_timestamp=now,
            )
        )
    for seed in surcharge_seeds:
        session.add(
            OtcSurcharge(
                product_type=seed.product_type, delivery_year=seed.delivery_year,
                surcharge_eur_mwh=seed.surcharge_eur_mwh,
            )
        )

    settlement_seeds, w = read_settlement_prices_from_excel(workbook, current_year)
    all_warnings += w
    settlement_date = today - dt.timedelta(days=1)
    for seed in settlement_seeds:
        session.add(
            SettlementPrice(
                product_type=seed.product_type, delivery_year=seed.delivery_year,
                settlement_date=settlement_date,
                settlement_price_eur_mwh=seed.settlement_price_eur_mwh,
                settlement_timestamp=now,
            )
        )

    calendar_seeds, w = read_trading_calendar_from_excel(workbook, current_year, today)
    all_warnings += w
    for seed in calendar_seeds:
        session.add(
            TradingCalendarEntry(
                due_date=seed.due_date, partner_alias=seed.partner_alias, direction=seed.direction,
                quantity_y0_mwh=seed.quantity_y0_mwh, quantity_y1_mwh=seed.quantity_y1_mwh,
                quantity_y2_mwh=seed.quantity_y2_mwh, quantity_y3_mwh=seed.quantity_y3_mwh,
                quantity_y4_mwh=seed.quantity_y4_mwh, status="geplant",
            )
        )

    limit_order_seeds, w = read_limit_orders_from_excel(workbook, current_year)
    all_warnings += w
    for seed in limit_order_seeds:
        session.add(
            LimitOrder(
                partner_alias=seed.partner_alias,
                quantity_y0_mwh=seed.quantity_y0_mwh, quantity_y1_mwh=seed.quantity_y1_mwh,
                quantity_y2_mwh=seed.quantity_y2_mwh, quantity_y3_mwh=seed.quantity_y3_mwh,
                quantity_y4_mwh=seed.quantity_y4_mwh,
                trigger_product_type=seed.trigger_product_type,
                trigger_delivery_year=seed.trigger_delivery_year,
                trigger_condition=seed.trigger_condition,
                limit_price_eur_mwh=seed.limit_price_eur_mwh,
                responsible_trading=seed.responsible_trading,
                responsible_sales=seed.responsible_sales,
                valid_until=seed.valid_until,
                status="offen",
            )
        )

    if pfc_dir is not None:
        pfc_seeds, w = read_pfc_checks_from_files(pfc_dir, current_year, today)
        all_warnings += w
        for seed in pfc_seeds:
            session.add(
                PfcCheck(
                    product_type=seed.product_type, delivery_year=seed.delivery_year,
                    pfc_mean_eur_mwh=seed.pfc_mean_eur_mwh, pfc_file_timestamp=seed.pfc_file_timestamp,
                )
            )

    upsert_metadata(session, "db_initialized", "true", now)
    upsert_metadata(session, "initial_seed_source", SOURCE_LABEL, now)
    if all_warnings:
        upsert_metadata(session, "last_import_warnings", " | ".join(all_warnings), now)

    return ImportReport(warnings=all_warnings, already_seeded=False)
